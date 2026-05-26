#!/usr/bin/env python3
"""
Adds the logged-in user's survey data (from Appwrite) as a new row in
data/teilnehmer.xlsx, then patches their Appwrite document with the
coordinates, ClaVis placeholder scores, and clavisId.

The Appwrite document $id is reused as the Excel row ID so that
FastAPI finds them via /matches/{doc.$id} without any frontend changes.

Usage:
  python3 scripts/add_user_to_excel.py
"""

import json
import shutil
import urllib.request
import urllib.error
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("❌  pip install openpyxl")
    raise


# ── Config ────────────────────────────────────────────────────────────────────

ROOT      = Path(__file__).parent.parent
ENV_FILE  = ROOT / ".env.local"
EXCEL_SRC = ROOT / "data" / "teilnehmer.xlsx"
# Also patch the algorithm repo copy if it exists alongside this repo
ALGO_EXCEL = ROOT.parent / "roommate-matching-nordakademie" / "data" / "teilnehmer.xlsx"

DOC_ID = "6a1226eb0017272a7004"   # Appwrite participant document $id

# Placeholder ClaVis CORE scores (user hasn't taken the ClaVis test yet).
# Derived heuristically from survey responses: high structure/gemeinschaft,
# homeoffice rhythm → moderate-high A/P, middle S/O.
CLAVIS_PLACEHOLDER = {
    "S":          50,
    "A":          60,
    "O":          55,
    "P":          65,
    "Sicherheit": 45,
    "Stimulation":50,
}

# Hamburg center coords (user's location is "Hamburg", no precise address given)
LOCATION_LAT = 53.5753
LOCATION_LON = 10.0153
RADIUS_KM    = 15


# ── Helpers ───────────────────────────────────────────────────────────────────

def load_env(path: Path) -> dict:
    env = {}
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            k, _, v = line.partition("=")
            env[k.strip()] = v.strip()
    return env


def appwrite_get(endpoint, project, api_key, path):
    headers = {"X-Appwrite-Project": project, "X-Appwrite-Key": api_key}
    req = urllib.request.Request(f"{endpoint}{path}", headers=headers)
    with urllib.request.urlopen(req) as r:
        return json.loads(r.read())


def appwrite_patch(endpoint, project, api_key, path, body):
    headers = {
        "Content-Type": "application/json",
        "X-Appwrite-Project": project,
        "X-Appwrite-Key": api_key,
    }
    data = json.dumps(body).encode()
    req = urllib.request.Request(f"{endpoint}{path}", data=data, headers=headers, method="PATCH")
    with urllib.request.urlopen(req) as r:
        return json.loads(r.read())


def to_smoking(val: str | None) -> str:
    """Convert German Appwrite values to English Excel values."""
    mapping = {"ja": "yes", "nein": "no", "draussen": "outside", "yes": "yes", "no": "no"}
    return mapping.get(str(val).lower().strip(), "no") if val else "no"


def to_bool_str(val) -> str:
    return "true" if val else "false"


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    env = load_env(ENV_FILE)
    endpoint   = env["NEXT_PUBLIC_APPWRITE_ENDPOINT"]
    project    = env["NEXT_PUBLIC_APPWRITE_PROJECT_ID"]
    db         = env["NEXT_PUBLIC_APPWRITE_DATABASE_ID"]
    coll       = env.get("NEXT_PUBLIC_PARTICIPANTS_COLLECTION_ID", "participants")
    api_key    = env["APPWRITE_API_KEY"]

    print(f"\n📥  Fetching Appwrite document {DOC_ID}…")
    doc = appwrite_get(endpoint, project, api_key,
                       f"/databases/{db}/collections/{coll}/documents/{DOC_ID}")
    print(f"   ✓ {doc['anzeigename']} (age={doc['age']}, gender={doc['gender']})\n")

    # ── Build the Frage-triplet mapping: kennung → antwort ──────────────────
    # Source the kennung→question_text from row 2 of the existing Excel so
    # we keep consistent question texts throughout the file.
    wb_src = openpyxl.load_workbook(EXCEL_SRC)
    ws_src = wb_src.active

    # Read question texts from row 2 (any existing row works; they're identical)
    question_texts: dict[str, str] = {}   # kennung → question text
    tan_question_text = ""
    for i in range(1, 56):
        base   = 42 + (i - 1) * 3          # col of Kennung (1-indexed)
        ken    = ws_src.cell(2, base).value
        frage  = ws_src.cell(2, base + 1).value
        if ken == "none":
            tan_question_text = str(frage or "")
        elif ken:
            question_texts[str(ken).strip()] = str(frage or "")

    # Map: kennung → answer value for the new user
    raucher     = to_smoking(doc.get("raucher"))
    raucher_tol = to_smoking(doc.get("raucher_toleranz"))

    # pets: "nein" / None → empty  |  "ja" (accept any) → "hund,katze"
    haustiere_raw = doc.get("haustiere") or ""
    haustiere = "" if haustiere_raw.lower() in ("nein", "no", "none", "") else haustiere_raw

    haustiere_tol_raw = doc.get("haustiere_toleranz") or ""
    if haustiere_tol_raw.lower() in ("ja", "yes"):
        haustiere_toleranz = "hund,katze"
    elif haustiere_tol_raw.lower() in ("nein", "no", "none", ""):
        haustiere_toleranz = ""
    else:
        haustiere_toleranz = haustiere_tol_raw

    user_answers: dict[str, object] = {
        "anzeigename":         doc.get("anzeigename"),
        "rolle":               doc.get("rolle"),
        "budget_min":          doc.get("budget_min"),
        "budget_max":          doc.get("budget_max"),
        "wohnort_lat":         LOCATION_LAT,
        "wohnort_lon":         LOCATION_LON,
        "radius_km":           RADIUS_KM,
        "einzug_fruehestens":  doc.get("einzug_fruehestens"),
        "einzug_spaetestens":  doc.get("einzug_spaetestens"),
        "raucher":             raucher,
        "raucher_toleranz":    raucher_tol,
        "haustiere":           haustiere,
        "haustiere_toleranz":  haustiere_toleranz,
        "nur_frauen_wg":       to_bool_str(doc.get("nur_frauen_wg")),
        "tagesrhythmus":       doc.get("tagesrhythmus"),
        "sauberkeit":          doc.get("sauberkeit"),
        "selbst_sauberkeit":   doc.get("selbst_sauberkeit"),
        "gemeinschaft":        doc.get("gemeinschaft"),
        "besucher":            doc.get("besucher"),
        "laerm_toleranz":      doc.get("laerm_toleranz"),
        "struktur":            doc.get("struktur"),
        "teilen":              doc.get("teilen"),
        **{f"pss_{i}":   doc.get(f"pss_{i}")   for i in range(1, 6)},
        **{f"rs_{i}":    doc.get(f"rs_{i}")    for i in range(1, 6)},
        **{f"ksa_{i}":   doc.get(f"ksa_{i}")   for i in range(1, 6)},
        **{f"ksdo_{i}":  doc.get(f"ksdo_{i}")  for i in range(1, 4)},
        **{f"agg_{i}":   doc.get(f"agg_{i}")   for i in range(1, 4)},
        **{f"imp_{i}":   doc.get(f"imp_{i}")   for i in range(1, 3)},
        **{f"neu_{i}":   doc.get(f"neu_{i}")   for i in range(1, 4)},
        **{f"ucla_{i}":  doc.get(f"ucla_{i}")  for i in range(1, 7)},
    }

    # ── Build new Excel row ──────────────────────────────────────────────────
    # Start with a full-width row of None, then fill known columns.
    n_cols   = ws_src.max_column
    new_row  = [None] * n_cols

    # Direct ClaVis columns (1-indexed → 0-indexed for list)
    new_row[0]  = DOC_ID                             # A: ID
    new_row[1]  = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S.000Z")  # B: Datum
    new_row[2]  = "nak_wohnraum_2026"                # C: Testspezifikation
    new_row[3]  = "nak_wohnraum_2026"                # D: Kampagne
    new_row[4]  = None                               # E: Kennung (empty)
    new_row[5]  = doc.get("age")                     # F: Alter
    new_row[6]  = doc.get("gender")                  # G: Geschlecht
    new_row[7]  = "de"                               # H: Sprache
    new_row[8]  = f"https://clavis1.20flow7.com/testengine/link/evaluation?token=SYN_{DOC_ID}"
    new_row[9]  = CLAVIS_PLACEHOLDER["S"]            # J: S
    new_row[10] = CLAVIS_PLACEHOLDER["A"]            # K: A
    new_row[11] = CLAVIS_PLACEHOLDER["O"]            # L: O
    new_row[12] = CLAVIS_PLACEHOLDER["P"]            # M: P
    new_row[13] = CLAVIS_PLACEHOLDER["Sicherheit"]   # N: Sicherheit
    new_row[14] = CLAVIS_PLACEHOLDER["Stimulation"]  # O: Stimulation
    new_row[15] = "SYN"                              # P: PSI-Typ (placeholder)
    new_row[16] = "ISTP"                             # Q: MBTI (placeholder)
    # Cols 17-40 (Bindung/Gestaltung/Leistung/Demotivation subscores): leave None

    # Frage triplets — keep kennung and question text from existing rows,
    # replace Antwort with user's actual answer.
    # Q01 (TAN, kennung="none"):
    base = 42  # col 42 (1-indexed), 0-indexed = 41
    new_row[41] = "none"              # Kennung
    new_row[42] = tan_question_text   # Frage (question text)
    new_row[43] = "99999"             # Antwort (synthetic TAN)

    for i in range(2, 56):           # Q02–Q55
        base_0   = 42 + (i - 1) * 3 - 1   # 0-indexed Kennung col
        frage_0  = base_0 + 1
        antwort_0 = base_0 + 2
        # Read kennung from existing row to maintain consistent ordering
        ken = ws_src.cell(2, base_0 + 1).value   # +1 to convert to 1-indexed
        if ken and ken != "none":
            new_row[base_0]   = ken
            new_row[frage_0]  = question_texts.get(str(ken).strip(), "")
            new_row[antwort_0] = user_answers.get(str(ken).strip())

    wb_src.close()

    # ── Append row to a copy of the Excel ───────────────────────────────────
    wb = openpyxl.load_workbook(EXCEL_SRC)
    ws = wb.active
    ws.append(new_row)
    new_row_num = ws.max_row

    # Save back to same path
    wb.save(EXCEL_SRC)
    wb.close()
    print(f"✅  Added row {new_row_num} to {EXCEL_SRC}")
    print(f"   ID (used as FastAPI participant_id): {DOC_ID}")

    # Also copy to the algorithm repo if it exists
    if ALGO_EXCEL.exists():
        shutil.copy2(EXCEL_SRC, ALGO_EXCEL)
        print(f"✅  Synced to algorithm repo: {ALGO_EXCEL}")

    # ── Patch Appwrite document ──────────────────────────────────────────────
    print(f"\n🔄  Patching Appwrite document {DOC_ID}…")
    patch_data = {
        "clavisId":   DOC_ID,
        "wohnort_lat": LOCATION_LAT,
        "wohnort_lon": LOCATION_LON,
        "radius_km":   RADIUS_KM,
        "clavis_s":         CLAVIS_PLACEHOLDER["S"],
        "clavis_a":         CLAVIS_PLACEHOLDER["A"],
        "clavis_o":         CLAVIS_PLACEHOLDER["O"],
        "clavis_p":         CLAVIS_PLACEHOLDER["P"],
        "clavis_sicherheit":  CLAVIS_PLACEHOLDER["Sicherheit"],
        "clavis_stimulation": CLAVIS_PLACEHOLDER["Stimulation"],
    }
    appwrite_patch(endpoint, project, api_key,
                   f"/databases/{db}/collections/{coll}/documents/{DOC_ID}",
                   {"data": patch_data})
    print("   ✓ clavisId, coordinates, and ClaVis placeholder scores saved.\n")

    print("🎉  Done! The user is now in the Excel dataset.")
    print(f"   Login:          http://localhost:3000/login")
    print(f"   Participant ID: {DOC_ID}")
    print(f"   FastAPI call:   GET /matches/{DOC_ID}\n")


if __name__ == "__main__":
    main()
