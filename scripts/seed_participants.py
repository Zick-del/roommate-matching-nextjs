#!/usr/bin/env python3
"""
Seed Appwrite with all 50 synthetic participants from data/teilnehmer.xlsx.

Each participant gets:
  - An Appwrite auth account: email = {tan}@nak-demo.de, password = Demo1234!
  - A participant document with documentId = ClaVis ID (all fields populated)
  - surveyDone=true, completedStep=7  →  they land directly on /results

The Appwrite document $id equals the ClaVis participant ID, so FastAPI can
look them up directly from the Excel when NEXT_PUBLIC_MOCK is disabled.

Usage:
  pip install openpyxl          # if not already installed
  python3 scripts/seed_participants.py

Prerequisites:
  - Appwrite running at NEXT_PUBLIC_APPWRITE_ENDPOINT
  - Collection created:  node scripts/setup-appwrite.mjs
  - .env.local with APPWRITE_API_KEY set
"""

import json
import sys
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("❌  openpyxl not found.  Run: pip install openpyxl")
    sys.exit(1)


# ── Config ────────────────────────────────────────────────────────────────────

ROOT = Path(__file__).parent.parent
ENV_FILE = ROOT / ".env.local"
EXCEL_FILE = ROOT / "data" / "teilnehmer.xlsx"
DEMO_PASSWORD = "Demo1234!"


# ── Load .env.local ───────────────────────────────────────────────────────────

def load_env(path: Path) -> dict:
    env = {}
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            key, _, val = line.partition("=")
            env[key.strip()] = val.strip()
    return env


if not ENV_FILE.exists():
    print(f"❌  .env.local not found at {ENV_FILE}")
    sys.exit(1)

env = load_env(ENV_FILE)
ENDPOINT   = env.get("NEXT_PUBLIC_APPWRITE_ENDPOINT", "http://localhost/v1")
PROJECT_ID = env["NEXT_PUBLIC_APPWRITE_PROJECT_ID"]
DATABASE_ID = env["NEXT_PUBLIC_APPWRITE_DATABASE_ID"]
COLLECTION_ID = env.get("NEXT_PUBLIC_PARTICIPANTS_COLLECTION_ID", "participants")
API_KEY    = env.get("APPWRITE_API_KEY", "")

if not API_KEY:
    print("❌  APPWRITE_API_KEY missing from .env.local")
    sys.exit(1)


# ── Appwrite REST helpers ─────────────────────────────────────────────────────

HEADERS = {
    "Content-Type": "application/json",
    "X-Appwrite-Project": PROJECT_ID,
    "X-Appwrite-Key": API_KEY,
}


def appwrite(method: str, path: str, body: dict | None = None) -> dict:
    url = f"{ENDPOINT}{path}"
    data = json.dumps(body).encode("utf-8") if body else None
    req = urllib.request.Request(url, data=data, headers=HEADERS, method=method)
    try:
        with urllib.request.urlopen(req) as resp:
            return json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        raw = e.read().decode("utf-8")
        err = json.loads(raw) if raw else {}
        raise RuntimeError(f"HTTP {e.code} {path}: {err.get('message', raw)}")


# ── Excel parsing ─────────────────────────────────────────────────────────────

def _int(v, default=None):
    if v is None:
        return default
    try:
        return int(v)
    except (ValueError, TypeError):
        return default


def _float(v, default=None):
    if v is None:
        return default
    try:
        return float(v)
    except (ValueError, TypeError):
        return default


def _bool(v, default=False):
    if isinstance(v, bool):
        return v
    if isinstance(v, str):
        return v.strip().lower() in ("true", "1", "yes", "ja")
    return default


def _str(v, default=None):
    if v is None:
        return default
    s = str(v).strip()
    return s if s else default


def coords_to_city(lat, lon) -> str:
    if lat is None or lon is None:
        return "Hamburg"
    lat, lon = float(lat), float(lon)
    if lat > 54.2:
        return "Kiel"
    if lat > 53.85:
        return "Norderstedt"
    if lat < 53.47:
        return "Hamburg-Harburg"
    if lon < 9.95:
        return "Hamburg-Altona"
    if lon > 10.15:
        return "Hamburg-Wandsbek"
    return "Hamburg"


def parse_row(ws, row: int) -> dict:
    """Parse one Excel row into a flat participant dict."""

    def cell(col: int):
        return ws.cell(row, col).value

    # ── Direct ClaVis columns ──────────────────────────────────────────────
    clavis_id        = _str(cell(1))   # ID
    age              = _int(cell(6))   # Alter
    gender           = _str(cell(7))   # Geschlecht  ("f" / "m" / "d")
    clavis_s         = _int(cell(10))  # S
    clavis_a         = _int(cell(11))  # A
    clavis_o         = _int(cell(12))  # O
    clavis_p         = _int(cell(13))  # P
    clavis_sicherheit   = _int(cell(14))  # Sicherheit
    clavis_stimulation  = _int(cell(15))  # Stimulation

    # ── Build kennung→antwort map from Frage triplets ─────────────────────
    # Column layout for question i (1-indexed):
    #   Kennung = col 42 + (i-1)*3       (field key, e.g. "anzeigename")
    #   Frage   = col 42 + (i-1)*3 + 1   (question text shown to participant)
    #   Antwort = col 42 + (i-1)*3 + 2   (participant's answer)
    kv: dict = {}
    tan = None
    for i in range(1, 56):
        base    = 42 + (i - 1) * 3
        kennung = ws.cell(row, base).value
        antwort = ws.cell(row, base + 2).value
        if kennung == "none":
            tan = _str(antwort)   # Q01: TAN from the study
        elif kennung:
            kv[str(kennung).strip()] = antwort

    def k(key, cast=None, default=None):
        v = kv.get(key, default)
        if v is None:
            return default
        return cast(v) if cast else v

    lat = _float(k("wohnort_lat"))
    lon = _float(k("wohnort_lon"))

    return {
        "clavisId": clavis_id,
        "tan":      tan,
        # Demographics
        "age":      age,
        "gender":   gender,
        # ClaVis CORE scores (read-only; populated by import, not by survey)
        "clavis_s":           clavis_s,
        "clavis_a":           clavis_a,
        "clavis_o":           clavis_o,
        "clavis_p":           clavis_p,
        "clavis_sicherheit":  clavis_sicherheit,
        "clavis_stimulation": clavis_stimulation,
        # Identity
        "anzeigename": _str(k("anzeigename")),
        "rolle":       _str(k("rolle")),
        # Location
        "wohnort_lat": lat,
        "wohnort_lon": lon,
        "radius_km":   _int(k("radius_km")),
        "location":    coords_to_city(lat, lon),
        # Hard constraints
        "budget_min":          _int(k("budget_min")),
        "budget_max":          _int(k("budget_max")),
        "einzug_fruehestens":  _str(k("einzug_fruehestens")),
        "einzug_spaetestens":  _str(k("einzug_spaetestens")),
        "raucher":             _str(k("raucher")),
        "raucher_toleranz":    _str(k("raucher_toleranz")),
        "haustiere":           _str(k("haustiere")) or "none",
        "haustiere_toleranz":  _str(k("haustiere_toleranz")) or "none",
        "nur_frauen_wg":       _bool(k("nur_frauen_wg")),
        # Lifestyle
        "tagesrhythmus":    _str(k("tagesrhythmus")),
        "sauberkeit":       _int(k("sauberkeit")),
        "selbst_sauberkeit": _int(k("selbst_sauberkeit")),
        "gemeinschaft":     _int(k("gemeinschaft")),
        "besucher":         _int(k("besucher")),
        "laerm_toleranz":   _int(k("laerm_toleranz")),
        "struktur":         _int(k("struktur")),
        "teilen":           _int(k("teilen")),
        # PSS-5
        **{f"pss_{i}": _int(k(f"pss_{i}")) for i in range(1, 6)},
        # RS-5
        **{f"rs_{i}": _int(k(f"rs_{i}")) for i in range(1, 6)},
        # KSA-5
        **{f"ksa_{i}": _int(k(f"ksa_{i}")) for i in range(1, 6)},
        # KSDO-3
        **{f"ksdo_{i}": _int(k(f"ksdo_{i}")) for i in range(1, 4)},
        # Aggression-3, Impulsivität-2, Neurotizismus-3
        **{f"agg_{i}": _int(k(f"agg_{i}")) for i in range(1, 4)},
        **{f"imp_{i}": _int(k(f"imp_{i}")) for i in range(1, 3)},
        **{f"neu_{i}": _int(k(f"neu_{i}")) for i in range(1, 4)},
        # UCLA-6
        **{f"ucla_{i}": _int(k(f"ucla_{i}")) for i in range(1, 7)},
        # Survey state
        "surveyDone":    True,
        "completedStep": 7,
    }


# ── Appwrite operations ───────────────────────────────────────────────────────

def create_or_find_user(email: str, name: str) -> str | None:
    """Create Appwrite auth user. Returns userId. Falls back to search on 409."""
    try:
        result = appwrite("POST", "/users", {
            "userId": "unique()",
            "email": email,
            "password": DEMO_PASSWORD,
            "name": name,
        })
        return result["$id"]
    except RuntimeError as e:
        if "409" not in str(e):
            print(f"      ⚠ user create failed: {e}")
            return None
    # User already exists — find by email via search
    try:
        encoded = urllib.parse.quote(email)
        result = appwrite("GET", f"/users?search={encoded}&limit=5")
        for user in result.get("users", []):
            if user.get("email", "").lower() == email.lower():
                return user["$id"]
    except RuntimeError as e:
        print(f"      ⚠ user search failed: {e}")
    return None


def upsert_document(doc_id: str, user_id: str, data: dict) -> bool:
    """Create or update participant document. Returns True on success."""
    # Strip None values — Appwrite rejects null for non-nullable attributes
    payload = {k: v for k, v in data.items() if v is not None}
    payload["userId"] = user_id

    try:
        appwrite(
            "POST",
            f"/databases/{DATABASE_ID}/collections/{COLLECTION_ID}/documents",
            {
                "documentId": doc_id,
                "data": payload,
                # Any authenticated user can read/update (demo only; production
                # should use per-user permissions: read("user:{userId}"))
                "permissions": ['read("users")', 'update("users")'],
            },
        )
        return True
    except RuntimeError as e:
        if "409" not in str(e):
            print(f"      ⚠ doc create failed: {e}")
            return False

    # Document exists — patch it
    try:
        appwrite(
            "PATCH",
            f"/databases/{DATABASE_ID}/collections/{COLLECTION_ID}/documents/{doc_id}",
            {"data": payload},
        )
        return True
    except RuntimeError as e:
        print(f"      ⚠ doc update failed: {e}")
        return False


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print(f"\n🌱  Seeding Appwrite participants from Excel")
    print(f"   Endpoint:   {ENDPOINT}")
    print(f"   Database:   {DATABASE_ID}")
    print(f"   Collection: {COLLECTION_ID}")
    print(f"   Excel:      {EXCEL_FILE}\n")

    if not EXCEL_FILE.exists():
        print(f"❌  Excel file not found: {EXCEL_FILE}")
        sys.exit(1)

    wb   = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)
    ws   = wb.active
    total = ws.max_row - 1  # minus header row
    print(f"Found {total} participants.\n")

    credentials: list[tuple] = []
    ok = 0
    fail = 0

    for row in range(2, ws.max_row + 1):
        p          = parse_row(ws, row)
        clavis_id  = p.get("clavisId")
        if not clavis_id:
            print(f"[{row-1:02d}] ⚠ no ClaVis ID, skipping")
            fail += 1
            continue

        anzeigename = p.get("anzeigename") or f"Teilnehmer {row - 1}"
        tan         = p.get("tan") or str(row - 1)
        email       = f"{tan}@nak-demo.de"

        print(f"[{row-1:02d}/{total}] {anzeigename:<20} ({clavis_id[:16]}…)")

        user_id = create_or_find_user(email, anzeigename)
        if not user_id:
            print(f"      ✗ could not create/find user")
            fail += 1
            continue
        print(f"      user  {user_id[:20]}…")

        # Use ClaVis ID as document $id so FastAPI can look up by participant_id
        doc_data = {k: v for k, v in p.items() if k not in ("tan",)}
        if upsert_document(clavis_id, user_id, doc_data):
            print(f"      doc   ✓")
            credentials.append((anzeigename, email, clavis_id))
            ok += 1
        else:
            fail += 1

    wb.close()

    width = 65
    print(f"\n{'=' * width}")
    print(f"  Seeded {ok}/{total} participants   ({fail} failed)")
    print(f"{'=' * width}")
    if credentials:
        print(f"\n  {'Name':<22} {'Email':<32} ClaVis ID")
        print(f"  {'-' * 62}")
        for name, email, cid in credentials:
            print(f"  {name:<22} {email:<32} {cid}")
        print(f"\n  All passwords:  {DEMO_PASSWORD}")
        print(f"  Login URL:      http://localhost:3000/login\n")


if __name__ == "__main__":
    main()
